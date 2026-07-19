import os
import time
import json
import urllib.request
import urllib.parse
import openpyxl

# ── CONFIGURAÇÕES ──────────────────────────────────────────────────
# Aba PRODUTOS do Criador de OPs vF.xlsm é a fonte única de cadastro de
# produto (mão única: a planilha manda, o Firebase só espelha — nunca editar
# produto no Firebase/produtos.html, só na planilha). Schema conforme
# GERADOR_OP_SPEC.md secao 3.1.
WATCH_FILE = r"C:\Users\gcall\Kuryos\Servidor Kuryos - Documentos\04. Design de Produtos e PCP\2. Ordens de Produção\Criador de OPs vF.xlsm"
SHEET_NAME = "PRODUTOS"
DB_URL = "https://prod-kuryos-default-rtdb.firebaseio.com"
STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "produtos_watcher_state.json")
# ───────────────────────────────────────────────────────────────────

# Cabeçalho esperado na linha 1 (uma coluna por campo, nessa ordem) — se a
# planilha mudar de layout, o script para e avisa em vez de importar dado
# na coluna errada silenciosamente.
EXPECTED_HEADER = [
    "cod_produto", "descricao", "cliente", "cod_cliente", "volume_nominal_ml",
    "pecas_por_caixa", "prazo_validade_meses", "overfill_pct", "perda_linha_pct",
    "perda_processo_pct", "densidade_granel", "capacidade_reator_l", "ativo",
]


def sanitize_key(val):
    s = str(val or '').replace('.', '-').replace('/', '-').replace('[', '-').replace(']', '-').replace('#', '-').replace('$', '-')
    s = '_'.join(s.split())
    return s[:60]


def load_state():
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, 'r') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_state(data):
    try:
        with open(STATE_FILE, 'w') as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        print("Erro ao salvar estado:", e)


def num_or_none(val):
    if val is None or val == '':
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def sync_produtos():
    if not os.path.exists(WATCH_FILE):
        print(f"  -> Arquivo não encontrado ainda: {WATCH_FILE}")
        return

    try:
        mtime = os.path.getmtime(WATCH_FILE)
    except Exception:
        return  # arquivo bloqueado/em gravação, tenta de novo no próximo ciclo

    state = load_state()
    if state.get("_last_mtime") == mtime:
        return  # nada mudou desde a última sincronização

    print(f"\nDetectada mudança em {os.path.basename(WATCH_FILE)}, sincronizando produtos...")
    try:
        wb = openpyxl.load_workbook(WATCH_FILE, data_only=True, read_only=True)
    except PermissionError:
        print("  -> Arquivo aberto no Excel no momento (locked). Tentando de novo no próximo ciclo.")
        return
    except Exception as e:
        print("  -> Erro ao abrir a planilha:", e)
        return

    if SHEET_NAME not in wb.sheetnames:
        print(f"  -> Aba '{SHEET_NAME}' não encontrada ainda. Abas disponíveis: {wb.sheetnames}")
        return

    sheet = wb[SHEET_NAME]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("  -> Aba PRODUTOS está vazia.")
        return

    header = [str(h).strip() if h else '' for h in rows[0]]
    header_ok = all(
        i < len(header) and header[i].strip().lower() == expected.lower()
        for i, expected in enumerate(EXPECTED_HEADER)
    )
    if not header_ok:
        print(f"  -> ATENÇÃO: cabeçalho da aba PRODUTOS não bate com o esperado.")
        print(f"     Esperado: {EXPECTED_HEADER}")
        print(f"     Encontrado: {header}")
        print("     Sincronização abortada até o layout ser corrigido — nada foi enviado ao Firebase.")
        return

    synced = 0
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        cod_produto = str(row[0]).strip()
        if not cod_produto:
            continue

        descricao = str(row[1]).strip() if row[1] else ""
        if not descricao:
            continue

        ativo_raw = row[12] if len(row) > 12 else None
        ativo = not (ativo_raw is False or str(ativo_raw or '').strip().lower() in ('false', 'inativo', 'não', 'nao', '0'))

        data = {
            "sku": cod_produto,
            "descricao": descricao,
            "cliente": str(row[2]).strip() if len(row) > 2 and row[2] else "",
            "codCliente": str(row[3]).strip() if len(row) > 3 and row[3] else "",
            "volume": num_or_none(row[4]) if len(row) > 4 else None,
            "unCx": num_or_none(row[5]) if len(row) > 5 else None,
            "prazoValidadeMeses": num_or_none(row[6]) if len(row) > 6 else None,
            "overfillPct": num_or_none(row[7]) if len(row) > 7 else None,
            "perdaLinhaPct": num_or_none(row[8]) if len(row) > 8 else None,
            "perdaProcessoPct": num_or_none(row[9]) if len(row) > 9 else None,
            "densidadeGranel": num_or_none(row[10]) if len(row) > 10 else None,
            "capacidadeReatorL": num_or_none(row[11]) if len(row) > 11 else None,
            "ativo": "Ativo" if ativo else "Inativo",
            "fonte": "produtos_watcher (Criador de OPs vF.xlsm)",
        }

        key = sanitize_key(cod_produto)
        url = f"{DB_URL}/produtos/{key}.json"

        # Preserva campos calculados só pelo Firebase (ex: prodHoraCalc) se existirem
        try:
            req = urllib.request.Request(url)
            with urllib.request.urlopen(req) as response:
                res = response.read()
                if res and res != b'null':
                    existing = json.loads(res.decode('utf-8'))
                    if existing.get("prodHoraCalc"):
                        data["prodHoraCalc"] = existing["prodHoraCalc"]
        except Exception:
            pass

        data_json = json.dumps(data).encode('utf-8')
        req = urllib.request.Request(url, data=data_json, method='PUT')
        req.add_header('Content-Type', 'application/json')
        try:
            with urllib.request.urlopen(req):
                synced += 1
        except Exception as e:
            print(f"  -> Falha ao sincronizar {cod_produto}:", e)

    print(f"  -> SUCESSO: {synced} produto(s) sincronizado(s) no Firebase!")
    state["_last_mtime"] = mtime
    save_state(state)


if __name__ == "__main__":
    print("==========================================================")
    print("   KURYOS PCP - WATCHER DE PRODUTOS (Criador de OPs vF)   ")
    print("==========================================================")
    print(f"Monitorando: {WATCH_FILE}")
    print(f"Aba: {SHEET_NAME}")
    print("Sincronização é MÃO ÚNICA: a planilha manda, o Firebase só espelha.")
    print("Pressione Ctrl+C para encerrar o monitoramento.")
    print("----------------------------------------------------------")
    while True:
        try:
            sync_produtos()
            time.sleep(10)  # varre a cada 10s
        except KeyboardInterrupt:
            print("\nMonitoramento encerrado pelo usuário.")
            break
        except Exception as e:
            print("Erro de varredura:", e)
            time.sleep(10)
